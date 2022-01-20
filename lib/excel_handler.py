import sys
import os
import openpyxl
import math
from openpyxl import drawing

def generate_output(data, template, output_file, mounting, level, drawing_grids, subsystem):

    # template setting
    starting_row = 10
    ending_row = 26
    fix_col = 1

    # template setting calculation
    row_count = ending_row - starting_row + 1
    data_count = len(data)
    page_num = math.ceil(data_count/row_count)

    if not os.path.isfile(template) : 
        print("\t\t\tError : template {} doest not exist".format(template))
        return

    wb = openpyxl.load_workbook(template)
    ws = wb.active

    # set up the first page 
    ws.title="Page_1"
    # page number
    ws['X5']="{}/{}".format(1,page_num)
    # area mounting
    ws['B6']="Area : {}".format(mounting)
    # system
    ws['B7']="{}".format(subsystem[:3])
    # drawing grids
    ws['B8']="{}".format(drawing_grids)
    # level
    ws['P6']="{}".format(level)
    # subsystem
    ws['P7']="{}".format(subsystem)
    insert_image(ws)

    read_row = 0

    # create page and fill in info
    for i in range(1, page_num):
        # copy the correct number of pages
        ws = wb.copy_worksheet(ws)
        insert_image(ws)
        ws.title = "Page_{}".format(i + 1)

        # update the page number
        ws['V5']="{}/{}".format(i + 1,page_num)

    # loop through each of the page
    for p in range(1, page_num + 1) : 
        # set correct page
        ws = wb['Page_' + str(p)]
        # insert every row in this page
        for i in range(0, row_count) :
            if read_row >= data_count :
                break
            ws.cell(row= starting_row + i, column= fix_col ).value= data[read_row]
            read_row += 1
    wb.save(output_file)

def insert_image(ws):
    img = drawing.image.Image('lib/tk-logo.PNG')
    ws.add_image(img, 'A2')
    img.width = 180
    img.height = 30
    img = drawing.image.Image('lib/bg-logo.PNG')
    ws.add_image(img, 'T1')
    img.width = 150
    img.height = 30
